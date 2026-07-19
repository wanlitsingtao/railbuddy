"""RailBuddy Web 管理服务器

基于 Flask 的轻量级 Web 管理界面，提供：
- 仪表盘：系统状态总览
- 数据源管理：网站/公众号的增删改查
- 邮箱配置：SMTP 设置和测试
- 调度配置：定时任务管理
- 抓取记录：分页浏览、搜索过滤
- 发送日志：历史发送记录
- 手动操作：立即抓取、测试邮件
"""

import os
import logging
import threading
import yaml
import tempfile
from datetime import datetime
from typing import Optional

from flask import Flask, request, jsonify, send_from_directory, render_template, make_response, send_file

from ..config import Config, ConfigError
from ..database import Database
from ..fetchers.registry import FetcherManager
from ..mailer import MailSender, get_provider_hint, EMAIL_PROVIDER_PRESETS
from ..models import BidItem, BidRecord

logger = logging.getLogger(__name__)


class WebServer:
    """RailBuddy Web 管理服务器"""

    def __init__(self, config_path: str, host: str = "0.0.0.0", port: int = 5210):
        self.config_path = os.path.abspath(config_path)
        self.base_dir = os.path.dirname(self.config_path)
        self.host = host
        self.port = port

        self.app = Flask(
            __name__,
            template_folder=os.path.join(os.path.dirname(__file__), "templates"),
            static_folder=os.path.join(os.path.dirname(__file__), "static"),
        )
        self._register_routes()

    def _resolve_path(self, path: str) -> str:
        if os.path.isabs(path):
            return path
        return os.path.join(self.base_dir, path)

    def _load_config(self) -> Config:
        """每次请求时重新加载配置，确保拿到最新值"""
        return Config(self.config_path)

    def _get_db(self, config: Config) -> Database:
        db_path = self._resolve_path(config.db_path)
        return Database(db_path, config.retention_days)

    def _register_routes(self):
        """注册所有路由"""
        app = self.app

        # ---- 页面路由 ----
        @app.route("/")
        def index():
            response = make_response(render_template("index.html"))
            response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
            return response

        # ---- 仪表盘 ----
        @app.route("/api/dashboard")
        def api_dashboard():
            try:
                config = self._load_config()
                db = self._get_db(config)
                stats = db.get_stats()
                source_states = db.get_all_source_states()
                last_send = db.get_last_send_log()
                send_logs = db.get_send_logs(10)

                # 配置摘要
                enabled_sources = len(config.sources)
                enabled_wechat = len(config.wechat_sources)
                total_sources = len(config.raw_sources) + len(config.raw_wechat_sources)
                schedule_times = config.schedule_config.get("times", [])

                # 仪表盘：数据源状态分页（只取首页 20 条 + 总数）
                source_states_result = db.get_source_states_paginated(page=1, per_page=20)
                recent_send_logs = db.get_send_logs_paginated(page=1, per_page=5)

                return jsonify({
                    "success": True,
                    "data": {
                        "stats": stats,
                        "source_states": source_states_result["items"],
                        "source_states_total": source_states_result["total"],
                        "last_send": last_send,
                        "send_logs": recent_send_logs["items"],
                        "config_summary": {
                            "total_sources": total_sources,
                            "enabled_sources": enabled_sources + enabled_wechat,
                            "disabled_sources": total_sources - enabled_sources - enabled_wechat,
                            "schedule_times": schedule_times,
                            "timezone": config.schedule_config.get("timezone", "Asia/Shanghai"),
                            "auto_send": config.schedule_config.get("auto_send", True),
                            "fetch_on_start": config.schedule_config.get("fetch_on_start", True),
                            "email_configured": bool(config.email_config.get("sender") and config.email_config.get("password")),
                            "receiver": config.email_config.get("receiver", ""),
                        }
                    }
                })
            except Exception as e:
                logger.error(f"Dashboard API error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 网站数据源 CRUD ----
        @app.route("/api/sources", methods=["GET"])
        def api_list_sources():
            try:
                config = self._load_config()
                sources = config.raw_sources
                page = int(request.args.get("page", 1))
                per_page = int(request.args.get("per_page", 20))
                total = len(sources)
                start = (page - 1) * per_page
                end = start + per_page
                page_sources = sources[start:end]
                return jsonify({
                    "success": True,
                    "data": {
                        "items": page_sources,
                        "total": total,
                        "page": page,
                        "per_page": per_page,
                        "total_pages": (total + per_page - 1) // per_page if per_page > 0 else 0
                    }
                })
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/sources", methods=["POST"])
        def api_add_source():
            try:
                source = request.json
                if not source or not source.get("name"):
                    return jsonify({"success": False, "error": "name 不能为空"}), 400

                src_type = source.get("type", "website")
                if src_type == "api":
                    if not source.get("api_url"):
                        return jsonify({"success": False, "error": "API 类型需要填写 api_url"}), 400
                else:
                    if not source.get("url"):
                        return jsonify({"success": False, "error": "网站类型需要填写 url"}), 400

                config = self._load_config()
                sources = config.raw_sources
                # 检查重名
                for s in sources:
                    if s["name"] == source["name"]:
                        return jsonify({"success": False, "error": f"数据源 '{source['name']}' 已存在"}), 400

                # 确保必要字段
                source.setdefault("type", "website")
                source.setdefault("enabled", True)
                source.setdefault("keywords", ["招标", "采购", "中标", "成交", "公告"])
                source.setdefault("fetch_detail", False)
                source.setdefault("timeout", 15)
                source.setdefault("request_interval", 2)

                sources.append(source)
                config.update_sources(sources)
                config.save_to_file()

                return jsonify({"success": True, "message": f"数据源 '{source['name']}' 已添加"})
            except Exception as e:
                logger.error(f"Add source error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/sources/<int:index>", methods=["PUT"])
        def api_update_source(index):
            try:
                config = self._load_config()
                sources = config.raw_sources
                if index < 0 or index >= len(sources):
                    return jsonify({"success": False, "error": "索引超出范围"}), 400

                source = request.json
                if not source or not source.get("name"):
                    return jsonify({"success": False, "error": "name 不能为空"}), 400

                src_type = source.get("type", "website")
                if src_type == "api":
                    if not source.get("api_url"):
                        return jsonify({"success": False, "error": "API 类型需要填写 api_url"}), 400
                else:
                    if not source.get("url"):
                        return jsonify({"success": False, "error": "网站类型需要填写 url"}), 400

                # 检查重名（排除自身）
                for i, s in enumerate(sources):
                    if i != index and s["name"] == source["name"]:
                        return jsonify({"success": False, "error": f"数据源 '{source['name']}' 已存在"}), 400

                sources[index] = source
                config.update_sources(sources)
                config.save_to_file()

                return jsonify({"success": True, "message": f"数据源 '{source['name']}' 已更新"})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/sources/<int:index>", methods=["DELETE"])
        def api_delete_source(index):
            try:
                config = self._load_config()
                sources = config.raw_sources
                if index < 0 or index >= len(sources):
                    return jsonify({"success": False, "error": "索引超出范围"}), 400

                name = sources[index].get("name", "")
                sources.pop(index)
                config.update_sources(sources)
                config.save_to_file()

                return jsonify({"success": True, "message": f"数据源 '{name}' 已删除"})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 公众号数据源 CRUD ----
        @app.route("/api/wechat-sources", methods=["GET"])
        def api_list_wechat():
            try:
                config = self._load_config()
                sources = config.raw_wechat_sources
                page = int(request.args.get("page", 1))
                per_page = int(request.args.get("per_page", 20))
                total = len(sources)
                start = (page - 1) * per_page
                end = start + per_page
                page_sources = sources[start:end]
                return jsonify({
                    "success": True,
                    "data": {
                        "items": page_sources,
                        "total": total,
                        "page": page,
                        "per_page": per_page,
                        "total_pages": (total + per_page - 1) // per_page if per_page > 0 else 0
                    }
                })
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/wechat-sources", methods=["POST"])
        def api_add_wechat():
            try:
                source = request.json
                if not source or not source.get("name"):
                    return jsonify({"success": False, "error": "name 不能为空"}), 400

                config = self._load_config()
                sources = config.raw_wechat_sources
                for s in sources:
                    if s["name"] == source["name"]:
                        return jsonify({"success": False, "error": f"公众号 '{source['name']}' 已存在"}), 400

                source.setdefault("type", "wechat")
                source.setdefault("enabled", False)
                source.setdefault("fetch_method", "sogou")
                source.setdefault("keywords", ["招标", "采购", "中标"])
                source.setdefault("timeout", 15)
                source.setdefault("request_interval", 3)

                sources.append(source)
                config.update_wechat_sources(sources)
                config.save_to_file()

                return jsonify({"success": True, "message": f"公众号 '{source['name']}' 已添加"})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/wechat-sources/<int:index>", methods=["PUT"])
        def api_update_wechat(index):
            try:
                config = self._load_config()
                sources = config.raw_wechat_sources
                if index < 0 or index >= len(sources):
                    return jsonify({"success": False, "error": "索引超出范围"}), 400

                source = request.json
                if not source or not source.get("name"):
                    return jsonify({"success": False, "error": "name 不能为空"}), 400

                for i, s in enumerate(sources):
                    if i != index and s["name"] == source["name"]:
                        return jsonify({"success": False, "error": f"公众号 '{source['name']}' 已存在"}), 400

                sources[index] = source
                config.update_wechat_sources(sources)
                config.save_to_file()

                return jsonify({"success": True, "message": f"公众号 '{source['name']}' 已更新"})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/wechat-sources/<int:index>", methods=["DELETE"])
        def api_delete_wechat(index):
            try:
                config = self._load_config()
                sources = config.raw_wechat_sources
                if index < 0 or index >= len(sources):
                    return jsonify({"success": False, "error": "索引超出范围"}), 400

                name = sources[index].get("name", "")
                sources.pop(index)
                config.update_wechat_sources(sources)
                config.save_to_file()

                return jsonify({"success": True, "message": f"公众号 '{name}' 已删除"})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 邮箱配置 ----
        @app.route("/api/email", methods=["GET"])
        def api_get_email():
            try:
                config = self._load_config()
                email_cfg = dict(config.email_config)
                # 脱敏：不返回密码，但保留 password_set 标志
                if email_cfg.get("password"):
                    email_cfg["password"] = ""
                    email_cfg["password_set"] = True
                else:
                    email_cfg["password_set"] = False
                return jsonify({"success": True, "data": email_cfg})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/email", methods=["PUT"])
        def api_update_email():
            try:
                config = self._load_config()
                email_cfg = request.json

                # 如果密码是占位符，保留原密码
                if email_cfg.get("password") == "******" or not email_cfg.get("password"):
                    old_config = config.email_config
                    email_cfg["password"] = old_config.get("password", "")

                config.update_email(email_cfg)
                config.save_to_file()

                return jsonify({"success": True, "message": "邮箱配置已更新"})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 调度配置 ----
        @app.route("/api/schedule", methods=["GET"])
        def api_get_schedule():
            try:
                config = self._load_config()
                return jsonify({"success": True, "data": config.schedule_config})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/schedule", methods=["PUT"])
        def api_update_schedule():
            try:
                config = self._load_config()
                sched_cfg = request.json
                if not sched_cfg.get("times"):
                    return jsonify({"success": False, "error": "times 不能为空"}), 400

                config.update_schedule(sched_cfg)
                config.save_to_file()

                return jsonify({"success": True, "message": "调度配置已更新"})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 去重配置 ----
        @app.route("/api/dedup", methods=["GET"])
        def api_get_dedup():
            try:
                config = self._load_config()
                return jsonify({"success": True, "data": config.dedup_config})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/dedup", methods=["PUT"])
        def api_update_dedup():
            try:
                config = self._load_config()
                dedup_cfg = request.json
                config.update_dedup(dedup_cfg)
                config.save_to_file()
                return jsonify({"success": True, "message": "去重配置已更新"})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 抓取记录 ----
        @app.route("/api/categories")
        def api_categories():
            """返回数据库中所有存在的类别列表"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                categories = db.get_all_categories()
                return jsonify({"success": True, "data": categories})
            except Exception as e:
                logger.error(f"Categories API error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 中标抓取分类管理（可配置的抓取条件） ----
        @app.route("/api/bid-categories", methods=["GET"])
        def api_bid_categories_list():
            """获取所有中标抓取分类配置"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                categories = db.get_bid_categories_config()
                return jsonify({"success": True, "data": categories})
            except Exception as e:
                logger.error(f"Bid categories list error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-categories", methods=["POST"])
        def api_bid_categories_add():
            """新增中标抓取分类"""
            try:
                data = request.json
                if not data or not data.get("name"):
                    return jsonify({"success": False, "error": "分类名称不能为空"}), 400
                config = self._load_config()
                db = self._get_db(config)
                cat_id = db.add_bid_category(
                    name=data["name"],
                    keywords=data.get("keywords", ""),
                    description=data.get("description", ""),
                    display_order=int(data.get("display_order", 0))
                )
                if cat_id:
                    return jsonify({"success": True, "data": {"id": cat_id}})
                return jsonify({"success": False, "error": "添加失败，可能名称已存在"}), 400
            except Exception as e:
                logger.error(f"Bid categories add error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-categories/<int:cat_id>", methods=["PUT"])
        def api_bid_categories_update(cat_id):
            """更新中标抓取分类"""
            try:
                data = request.json
                if not data:
                    return jsonify({"success": False, "error": "请求数据不能为空"}), 400
                config = self._load_config()
                db = self._get_db(config)
                ok = db.update_bid_category(cat_id, data)
                if ok:
                    return jsonify({"success": True})
                return jsonify({"success": False, "error": "更新失败"}), 400
            except Exception as e:
                logger.error(f"Bid categories update error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-categories/<int:cat_id>", methods=["DELETE"])
        def api_bid_categories_delete(cat_id):
            """删除中标抓取分类"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                ok = db.delete_bid_category(cat_id)
                if ok:
                    return jsonify({"success": True})
                return jsonify({"success": False, "error": "删除失败"}), 400
            except Exception as e:
                logger.error(f"Bid categories delete error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/items")
        def api_items():
            try:
                config = self._load_config()
                db = self._get_db(config)

                page = int(request.args.get("page", 1))
                per_page = int(request.args.get("per_page", 20))
                source = request.args.get("source") or None
                status = request.args.get("status") or None
                category = request.args.get("category") or None
                keyword = request.args.get("keyword") or None
                date_from = request.args.get("date_from") or None
                date_to = request.args.get("date_to") or None

                result = db.get_items_paginated(
                    page=page, per_page=per_page,
                    source=source, status=status, category=category,
                    keyword=keyword, date_from=date_from, date_to=date_to
                )
                sources_list = db.get_all_sources()

                return jsonify({
                    "success": True,
                    "data": result,
                    "sources": sources_list
                })
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 里程统计 ----
        @app.route("/api/transit-stats")
        def api_transit_stats():
            """里程数据统计概要"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                stats = db.get_mileage_stats()
                return jsonify({"success": True, "data": stats})
            except Exception as e:
                logger.error(f"Transit stats API error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/transit-trend")
        def api_transit_trend():
            """全国里程月度变化趋势"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                trend = db.get_national_trend()
                return jsonify({"success": True, "data": trend})
            except Exception as e:
                logger.error(f"Transit trend API error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/transit-cities")
        def api_transit_cities():
            """城市维度里程汇总（最新月）+ 城市维度趋势"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                city = request.args.get("city") or None
                # 城市排名（最新月）
                city_summary = db.get_city_summary()
                # 城市趋势
                city_trend = db.get_city_trend(city)
                # 可选城市列表
                cities = db.get_mileage_cities()
                return jsonify({
                    "success": True,
                    "data": city_summary,
                    "trend": city_trend,
                    "cities": cities,
                })
            except Exception as e:
                logger.error(f"Transit cities API error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/transit-mileage")
        def api_transit_mileage():
            """查询单月里程明细"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                month = request.args.get("month")
                if not month:
                    month = db.get_latest_mileage_month()
                if not month:
                    return jsonify({"success": True, "data": []})
                mileage = db.get_mileage_by_month(month)
                months = db.get_mileage_months()
                return jsonify({
                    "success": True,
                    "data": mileage,
                    "month": month,
                    "available_months": months,
                })
            except Exception as e:
                logger.error(f"Transit mileage API error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/transit-annual-trend")
        def api_transit_annual_trend():
            """全国年度里程趋势（1969-至今），含制式分类"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                trend = db.get_national_annual_trend()
                return jsonify({"success": True, "data": trend})
            except Exception as e:
                logger.error(f"Transit annual trend API error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/transit-system-types")
        def api_transit_system_types():
            """按制式分类的里程统计（最新月）"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                breakdown = db.get_system_type_breakdown()
                return jsonify({"success": True, "data": breakdown})
            except Exception as e:
                logger.error(f"Transit system types API error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/transit-city-trend")
        def api_transit_city_trend():
            """城市维度年度里程趋势"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                city = request.args.get("city", None)
                trend = db.get_city_annual_trend()
                if city:
                    trend = [t for t in trend if t["city"] == city]
                return jsonify({"success": True, "data": trend})
            except Exception as e:
                logger.error(f"Transit city trend API error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/seed-history", methods=["POST"])
        def api_seed_history():
            """加载历史里程数据"""
            try:
                from ..data.transit_history import seed_history
                config = self._load_config()
                db = self._get_db(config)
                result = seed_history(db)
                return jsonify({"success": True, "data": result})
            except Exception as e:
                logger.error(f"Seed history API error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 发送日志 ----
        @app.route("/api/send-logs")
        def api_send_logs():
            try:
                config = self._load_config()
                db = self._get_db(config)
                page = int(request.args.get("page", 1))
                per_page = int(request.args.get("per_page", 20))
                result = db.get_send_logs_paginated(page=page, per_page=per_page)
                return jsonify({"success": True, "data": result})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 数据源状态 ----
        @app.route("/api/source-states")
        def api_source_states():
            try:
                config = self._load_config()
                db = self._get_db(config)
                page = int(request.args.get("page", 1))
                per_page = int(request.args.get("per_page", 20))
                result = db.get_source_states_paginated(page=page, per_page=per_page)
                return jsonify({"success": True, "data": result})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 手动操作：立即抓取 ----
        @app.route("/api/fetch-now", methods=["POST"])
        def api_fetch_now():
            try:
                config = self._load_config()
                db = self._get_db(config)

                # 支持时间范围参数：信息发布时间区间
                data = request.json or {}
                date_from = data.get("date_from") or None
                date_to = data.get("date_to") or None

                fetcher_mgr = FetcherManager(
                    sources_config=config.sources,
                    wechat_sources_config=config.wechat_sources,
                    max_items_per_source=config.max_items_per_source,
                    max_age_days=0  # 手动抓取不限制历史天数
                )
                new_items = fetcher_mgr.fetch_all(db, date_from=date_from, date_to=date_to)

                # 检查是否启用自动发送
                auto_send = config.schedule_config.get("auto_send", True)

                sent_count = 0
                send_status = "skipped"
                if new_items and auto_send:
                    mailer = MailSender(config.email_config)
                    success = mailer.send(new_items)
                    if success:
                        item_ids = [item.item_id for item in new_items]
                        db.mark_items_sent(item_ids)
                        sources_involved = set(item.source for item in new_items)
                        for sn in sources_involved:
                            db.update_send_time(sn)
                        sent_count = len(new_items)
                        send_status = "success"
                        db.log_send(sent_count, config.email_config.get("receiver", ""), "success")
                    else:
                        send_status = "failed"
                        db.log_send(len(new_items), config.email_config.get("receiver", ""), "failed", "SMTP 发送失败")
                elif not auto_send:
                    send_status = "auto_send_disabled"
                else:
                    # 检查是否有遗留未发送的
                    pending = db.get_unsent_items()
                    if pending:
                        mailer = MailSender(config.email_config)
                        success = mailer.send(pending)
                        if success:
                            item_ids = [item.item_id for item in pending]
                            db.mark_items_sent(item_ids)
                            sent_count = len(pending)
                            send_status = "success_retry"
                            db.log_send(sent_count, config.email_config.get("receiver", ""), "success")

                return jsonify({
                    "success": True,
                    "data": {
                        "new_items": len(new_items),
                        "sent": sent_count,
                        "send_status": send_status
                    }
                })
            except Exception as e:
                logger.error(f"Fetch now error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 邮箱提供商预设 ----
        @app.route("/api/email-providers", methods=["GET"])
        def api_email_providers():
            """返回常见邮箱提供商的 SMTP 预设配置"""
            return jsonify({"success": True, "data": EMAIL_PROVIDER_PRESETS})

        # ---- 手动操作：SMTP 诊断 ----
        @app.route("/api/diagnose-smtp", methods=["POST"])
        def api_diagnose_smtp():
            """诊断 SMTP 连接问题，返回详细的步骤级诊断结果"""
            try:
                config = self._load_config()
                mailer = MailSender(config.email_config)
                ok, detail = mailer.diagnose()
                return jsonify({
                    "success": True,
                    "data": {
                        "connected": ok,
                        "detail": detail,
                        "config": {
                            "server": config.email_config.get("smtp_server", ""),
                            "port": config.email_config.get("smtp_port", 465),
                            "use_ssl": config.email_config.get("use_ssl", True),
                            "sender": config.email_config.get("sender", ""),
                        }
                    }
                })
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 手动操作：测试邮件 ----
        @app.route("/api/test-email", methods=["POST"])
        def api_test_email():
            try:
                config = self._load_config()
                mailer = MailSender(config.email_config)
                test_item = BidItem(
                    title="[测试] RailBuddy 邮件配置验证",
                    url="https://example.com",
                    source="RailBuddy 测试",
                    publish_date=datetime.now().strftime("%Y-%m-%d"),
                    description="如果您收到了这封邮件，说明邮箱配置正确。",
                    category="测试"
                )
                success = mailer.send([test_item])
                if success:
                    return jsonify({"success": True, "message": "测试邮件已发送，请检查收件箱"})
                else:
                    # 发送失败时附带诊断信息
                    ok, detail = mailer.diagnose()
                    provider_hint = get_provider_hint(config.email_config.get("sender", ""))
                    hint_text = ""
                    if provider_hint:
                        hint_text = f"\n{provider_hint['label']}提示：{provider_hint['note']}"
                    return jsonify({
                        "success": False,
                        "error": "邮件发送失败",
                        "detail": detail + hint_text
                    }), 500
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 手动操作：发送选中条目 ----
        @app.route("/api/send-items", methods=["POST"])
        def api_send_items():
            try:
                data = request.json or {}
                item_ids = data.get("item_ids", [])
                if not item_ids:
                    return jsonify({"success": False, "error": "未选择任何条目"}), 400

                config = self._load_config()
                db = self._get_db(config)

                # 按 ID 查询条目
                items = db.get_items_by_ids(item_ids)
                if not items:
                    return jsonify({"success": False, "error": "未找到选中的条目"}), 404

                # 发送邮件
                mailer = MailSender(config.email_config)
                success = mailer.send(items)

                if success:
                    db.mark_items_sent(item_ids)
                    sources_involved = set(item.source for item in items)
                    for sn in sources_involved:
                        db.update_send_time(sn)
                    db.log_send(
                        len(items),
                        config.email_config.get("receiver", ""),
                        "success"
                    )
                    return jsonify({
                        "success": True,
                        "message": f"已发送 {len(items)} 条记录到邮箱",
                        "data": {"sent": len(items)}
                    })
                else:
                    db.log_send(
                        len(items),
                        config.email_config.get("receiver", ""),
                        "failed",
                        "SMTP 发送失败"
                    )
                    return jsonify({"success": False, "error": "邮件发送失败，请检查 SMTP 配置"}), 500
            except Exception as e:
                logger.error(f"Send items error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 项目中标记录 CRUD ----
        @app.route("/api/bid-records", methods=["GET"])
        def api_bid_records_list():
            """分页查询中标记录，支持多维度筛选"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                page = int(request.args.get("page", 1))
                per_page = int(request.args.get("per_page", 20))
                # 分类：支持逗号分隔的多选
                categories_str = request.args.get("categories") or None
                categories = categories_str.split(",") if categories_str else None
                category = request.args.get("category") or None  # 兼容旧的单选
                city = request.args.get("city") or None
                province = request.args.get("province") or None
                winner = request.args.get("winner") or None
                keyword = request.args.get("keyword") or None
                date_from = request.args.get("date_from") or None
                date_to = request.args.get("date_to") or None

                result = db.get_bid_records_paginated(
                    page=page, per_page=per_page,
                    categories=categories, category=category,
                    city=city, province=province,
                    winner=winner, keyword=keyword,
                    date_from=date_from, date_to=date_to
                )

                # 附加筛选选项列表
                categories_list = db.get_bid_categories()
                cities = db.get_bid_cities()
                provinces = db.get_bid_provinces()

                return jsonify({
                    "success": True,
                    "data": result,
                    "categories": categories_list,
                    "cities": cities,
                    "provinces": provinces,
                })
            except Exception as e:
                logger.error(f"Bid records list error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-records/<record_id>", methods=["GET"])
        def api_bid_record_detail(record_id):
            """获取单条中标记录详情"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                record = db.get_bid_record(record_id)
                if not record:
                    return jsonify({"success": False, "error": "记录不存在"}), 404
                return jsonify({"success": True, "data": record})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-records", methods=["POST"])
        def api_bid_record_create():
            """新增中标记录"""
            try:
                data = request.json
                if not data or not data.get("project_name"):
                    return jsonify({"success": False, "error": "项目名称不能为空"}), 400

                # 构建 BidRecord，record_id 自动生成
                record = BidRecord(
                    project_name=data["project_name"],
                    province=data.get("province", ""),
                    city=data.get("city", ""),
                    category=data.get("category", ""),
                    winner=data.get("winner", ""),
                    consortium=data.get("consortium", ""),
                    project_overview=data.get("project_overview", ""),
                    bid_scope=data.get("bid_scope", ""),
                    subsystems=data.get("subsystems", ""),
                    bid_threshold=data.get("bid_threshold", ""),
                    bidder=data.get("bidder", ""),
                    funding_source=data.get("funding_source", ""),
                    evaluation_method=data.get("evaluation_method", ""),
                    total_stations=data.get("total_stations"),
                    underground_stations=data.get("underground_stations"),
                    elevated_stations=data.get("elevated_stations"),
                    ground_stations=data.get("ground_stations"),
                    opened_stations=data.get("opened_stations"),
                    line_type=data.get("line_type", ""),
                    length_km=data.get("length_km"),
                    goa_level=data.get("goa_level", ""),
                    system_mode=data.get("system_mode", ""),
                    is_opened=data.get("is_opened", ""),
                    opening_date=data.get("opening_date"),
                    bid_date=data.get("bid_date"),
                    bid_amount=data.get("bid_amount"),
                    control_price=data.get("control_price"),
                    bid_link=data.get("bid_link", ""),
                    tender_link=data.get("tender_link", ""),
                    design_unit=data.get("design_unit", ""),
                    platform_software_pis=data.get("platform_software_pis", ""),
                    notes=data.get("notes", ""),
                    data_source="manual",
                )

                config = self._load_config()
                db = self._get_db(config)
                db.save_bid_record(record)

                return jsonify({
                    "success": True,
                    "message": "中标记录已创建",
                    "data": {"record_id": record.record_id}
                })
            except Exception as e:
                logger.error(f"Bid record create error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-records/<record_id>", methods=["PUT"])
        def api_bid_record_update(record_id):
            """更新中标记录（全字段更新）"""
            try:
                data = request.json
                if not data:
                    return jsonify({"success": False, "error": "无更新数据"}), 400

                config = self._load_config()
                db = self._get_db(config)

                # 不允许通过 API 修改 record_id
                data.pop("record_id", None)

                # 如果修改了 project_name 或 bid_date，需重新计算 record_id
                # 后端 update_bid_record 中处理重新生成 record_id
                success = db.update_bid_record(record_id, data)
                if not success:
                    return jsonify({"success": False, "error": "记录不存在"}), 404

                return jsonify({"success": True, "message": "中标记录已更新"})
            except Exception as e:
                logger.error(f"Bid record update error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-records/<record_id>", methods=["DELETE"])
        def api_bid_record_delete(record_id):
            """删除中标记录"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                success = db.delete_bid_record(record_id)
                if not success:
                    return jsonify({"success": False, "error": "记录不存在"}), 404

                return jsonify({"success": True, "message": "中标记录已删除"})
            except Exception as e:
                logger.error(f"Bid record delete error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-records/stats")
        def api_bid_records_stats():
            """中标记录统计概要，支持筛选条件动态计算"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                # 支持筛选参数
                categories_str = request.args.get("categories") or None
                categories = categories_str.split(",") if categories_str else None
                category = request.args.get("category") or None
                city = request.args.get("city") or None
                province = request.args.get("province") or None
                keyword = request.args.get("keyword") or None
                date_from = request.args.get("date_from") or None
                date_to = request.args.get("date_to") or None
                stats = db.get_bid_stats_filtered(
                    categories=categories, category=category,
                    city=city, province=province,
                    keyword=keyword, date_from=date_from, date_to=date_to
                )
                return jsonify({"success": True, "data": stats})
            except Exception as e:
                logger.error(f"Bid records stats error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-records/export")
        def api_bid_records_export():
            """导出中标记录到Excel"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                # 支持与查询相同的筛选参数
                categories_str = request.args.get("categories") or None
                categories = categories_str.split(",") if categories_str else None
                category = request.args.get("category") or None
                city = request.args.get("city") or None
                province = request.args.get("province") or None
                keyword = request.args.get("keyword") or None
                date_from = request.args.get("date_from") or None
                date_to = request.args.get("date_to") or None

                # 获取全部符合条件的记录（不分页）
                result = db.get_bid_records_paginated(
                    page=1, per_page=99999,
                    categories=categories, category=category,
                    city=city, province=province,
                    keyword=keyword, date_from=date_from, date_to=date_to
                )

                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "中标记录"

                # 表头
                headers = ["项目名称", "工程分类", "省份", "城市", "中标单位", "联合体",
                           "中标金额（万元）", "控制价（万元）", "中标日期", "开通状态",
                           "线路类型", "线路长度(km)", "站点数",
                           "GOA等级", "制式模式", "项目概况", "招标范围",
                           "备注", "招标链接"]
                ws.append(headers)

                # 数据行
                for item in result["items"]:
                    row = [
                        item.get("project_name", ""),
                        item.get("category", ""),
                        item.get("province", ""),
                        item.get("city", ""),
                        item.get("winner", ""),
                        item.get("consortium", ""),
                        item.get("bid_amount") or "",
                        item.get("control_price") or "",
                        item.get("bid_date", ""),
                        item.get("is_opened", ""),
                        item.get("line_type", ""),
                        item.get("length_km") or "",
                        item.get("total_stations") or "",
                        item.get("goa_level", ""),
                        item.get("system_mode", ""),
                        item.get("project_overview", ""),
                        item.get("bid_scope", ""),
                        item.get("notes", ""),
                        item.get("bid_link", ""),
                    ]
                    ws.append(row)

                # 保存到临时文件
                tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                wb.save(tmp.name)
                tmp.close()

                return send_file(
                    tmp.name,
                    as_attachment=True,
                    download_name="中标记录_" + datetime.now().strftime("%Y%m%d") + ".xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                logger.error(f"Bid records export error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-records/import-template")
        def api_bid_records_import_template():
            """下载中标记录导出模版"""
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "中标记录导出模版"

                # 模版表头 + 示例数据说明
                headers = ["项目名称*", "工程分类", "省份", "城市", "中标单位", "联合体",
                           "中标金额（万元）", "控制价（万元）", "中标日期*", "开通状态",
                           "线路类型", "线路长度(km)", "站点数",
                           "GOA等级", "制式模式", "项目概况", "招标范围",
                           "备注", "招标链接"]
                ws.append(headers)

                # 示例行
                ws.append([
                    "XX市X号线信号系统采购项目",
                    "信号",
                    "广东",
                    "深圳",
                    "交控科技股份有限公司",
                    "",
                    25800,
                    30000,
                    "2024-01-15",
                    "在建",
                    "地铁",
                    36.5,
                    25,
                    "GOA4",
                    "CBTC",
                    "XX号线信号系统...",
                    "信号系统设计、供货、安装...",
                    "专用通信",
                    "https://..."
                ])

                # 添加说明sheet
                ws2 = wb.create_sheet("填写说明")
                ws2.append(["字段名", "是否必填", "说明"])
                ws2.append(["项目名称*", "必填", "项目全称，与中标日期一起作为唯一标识"])
                ws2.append(["工程分类", "可选", "可选值：通信、综合监控、信号、安防、消防、线网、弱电、大总包、施工总包、BOT、PPP、其他"])
                ws2.append(["省份", "可选", "省份名称"])
                ws2.append(["城市", "可选", "城市名称"])
                ws2.append(["中标单位", "可选", "中标单位名称"])
                ws2.append(["联合体", "可选", "联合体成员"])
                ws2.append(["中标金额（万元）", "可选", "金额单位为万元，填写数值"])
                ws2.append(["控制价（万元）", "可选", "金额单位为万元，填写数值"])
                ws2.append(["中标日期*", "必填", "格式：YYYY-MM-DD"])
                ws2.append(["开通状态", "可选", "可选值：已开通、在建、未开通"])
                ws2.append(["线路类型", "可选", "如：地铁、轻轨、单轨等"])
                ws2.append(["线路长度(km)", "可选", "数值"])
                ws2.append(["站点数", "可选", "整数"])
                ws2.append(["GOA等级", "可选", "可选值：GOA1、GOA2、GOA3、GOA4"])
                ws2.append(["制式模式", "可选", "如：CBTC、互联互通"])
                ws2.append(["项目概况", "可选", "项目描述"])
                ws2.append(["招标范围", "可选", "招标范围描述"])
                ws2.append(["备注", "可选", "备注信息"])
                ws2.append(["招标链接", "可选", "招标公告URL"])

                tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                wb.save(tmp.name)
                tmp.close()

                return send_file(
                    tmp.name,
                    as_attachment=True,
                    download_name="中标记录导出模版.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                logger.error(f"Import template error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-records/import-upload", methods=["POST"])
        def api_bid_records_import_upload():
            """从用户上传的Excel文件导入中标记录
            支持 duplicate_action 参数: all(默认)/skip(跳过重复)/overwrite(覆盖重复)
            """
            try:
                if 'file' not in request.files:
                    return jsonify({"success": False, "error": "未找到上传文件"}), 400
                file = request.files['file']
                if not file.filename.endswith(('.xlsx', '.xls')):
                    return jsonify({"success": False, "error": "仅支持 .xlsx 或 .xls 文件"}), 400

                duplicate_action = request.args.get("duplicate_action", "all")

                # 保存上传文件到临时目录
                tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                file.save(tmp.name)
                tmp.close()

                config = self._load_config()
                db = self._get_db(config)

                from ..data.import_from_template import import_from_template
                result = import_from_template(tmp.name, db,
                                              duplicate_action=duplicate_action)

                # 清理临时文件
                try:
                    os.unlink(tmp.name)
                except Exception:
                    pass

                return jsonify({"success": True, "data": result})
            except Exception as e:
                logger.error(f"Bid records import upload error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-records/import-check", methods=["POST"])
        def api_bid_records_import_check():
            """上传Excel并检查重复（仅分析不导入）"""
            try:
                if 'file' not in request.files:
                    return jsonify({"success": False, "error": "未找到上传文件"}), 400
                file = request.files['file']
                if not file.filename.endswith(('.xlsx', '.xls')):
                    return jsonify({"success": False, "error": "仅支持 .xlsx 或 .xls 文件"}), 400

                tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                file.save(tmp.name)
                tmp.close()

                config = self._load_config()
                db = self._get_db(config)

                from ..data.import_from_template import import_from_template
                result = import_from_template(tmp.name, db, check_only=True)

                try:
                    os.unlink(tmp.name)
                except Exception:
                    pass

                return jsonify({"success": True, "data": result})
            except Exception as e:
                logger.error(f"Bid records import check error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 金额补全：搜索全网数据补全空缺金额 ----
        @app.route("/api/bid-records/fill-amounts/preview", methods=["POST"])
        def api_bid_fill_preview():
            """预览：列出所有存在金额空缺的记录"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                from ..data.fill_amounts import AmountFiller
                filler = AmountFiller(db)
                result = filler.preview()
                return jsonify({"success": True, "data": result})
            except Exception as e:
                logger.error(f"Fill amounts preview error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-records/fill-amounts", methods=["POST"])
        def api_bid_fill_execute():
            """执行金额补全：搜索全网数据填充空缺的金额"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                data = request.json or {}
                max_records = int(data.get("max_records", 10))

                from ..data.fill_amounts import AmountFiller
                filler = AmountFiller(db)

                # 在线程中执行（避免Web请求超时），但同时返回初始预览
                result_container = {"result": None, "error": None}

                def _do_fill():
                    try:
                        result_container["result"] = filler.fill_missing(max_records=max_records)
                    except Exception as ex:
                        result_container["error"] = str(ex)

                t = threading.Thread(target=_do_fill, daemon=True)
                t.start()
                t.join(timeout=300)  # 最多等5分钟

                if result_container["error"]:
                    return jsonify({"success": False, "error": result_container["error"]}), 500
                if result_container["result"] is None:
                    return jsonify({
                        "success": True,
                        "data": {"status": "running", "message": "补全任务正在后台执行，请稍后查看结果"}
                    })

                return jsonify({"success": True, "data": result_container["result"]})
            except Exception as e:
                logger.error(f"Fill amounts execute error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-records/fill-amounts/<record_id>", methods=["POST"])
        def api_bid_fill_single(record_id):
            """单条记录金额补全"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                from ..data.fill_amounts import AmountFiller
                filler = AmountFiller(db)
                result = filler.fill_single(record_id)
                return jsonify({"success": True, "data": result})
            except Exception as e:
                logger.error(f"Fill single amount error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        # ---- 中标动态（bid_raw）CRUD ----
        @app.route("/api/bid-dynamic", methods=["GET"])
        def api_bid_dynamic_list():
            """分页查询中标动态（bid_raw），支持多维度筛选"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                page = int(request.args.get("page", 1))
                per_page = int(request.args.get("per_page", 20))
                categories_str = request.args.get("categories") or None
                categories = categories_str.split(",") if categories_str else None
                category = request.args.get("category") or None
                city = request.args.get("city") or None
                province = request.args.get("province") or None
                winner = request.args.get("winner") or None
                keyword = request.args.get("keyword") or None
                date_from = request.args.get("date_from") or None
                date_to = request.args.get("date_to") or None

                result = db.get_bid_raw_paginated(
                    page=page, per_page=per_page,
                    categories=categories, category=category,
                    city=city, province=province,
                    winner=winner, keyword=keyword,
                    date_from=date_from, date_to=date_to
                )

                categories_list = db.get_bid_raw_categories()
                cities = db.get_bid_raw_cities()
                provinces = db.get_bid_raw_provinces()

                return jsonify({
                    "success": True,
                    "data": result,
                    "categories": categories_list,
                    "cities": cities,
                    "provinces": provinces,
                })
            except Exception as e:
                logger.error(f"Bid dynamic list error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-dynamic/<record_id>", methods=["GET"])
        def api_bid_dynamic_detail(record_id):
            """获取单条中标动态详情"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                record = db.get_bid_raw(record_id)
                if not record:
                    return jsonify({"success": False, "error": "记录不存在"}), 404
                return jsonify({"success": True, "data": record})
            except Exception as e:
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-dynamic", methods=["POST"])
        def api_bid_dynamic_create():
            """新增中标动态记录"""
            try:
                data = request.json
                if not data or not data.get("project_name"):
                    return jsonify({"success": False, "error": "项目名称不能为空"}), 400

                record = BidRecord(
                    project_name=data["project_name"],
                    province=data.get("province", ""),
                    city=data.get("city", ""),
                    category=data.get("category", ""),
                    winner=data.get("winner", ""),
                    consortium=data.get("consortium", ""),
                    project_overview=data.get("project_overview", ""),
                    bid_scope=data.get("bid_scope", ""),
                    subsystems=data.get("subsystems", ""),
                    bid_threshold=data.get("bid_threshold", ""),
                    bidder=data.get("bidder", ""),
                    funding_source=data.get("funding_source", ""),
                    evaluation_method=data.get("evaluation_method", ""),
                    total_stations=data.get("total_stations"),
                    underground_stations=data.get("underground_stations"),
                    elevated_stations=data.get("elevated_stations"),
                    ground_stations=data.get("ground_stations"),
                    opened_stations=data.get("opened_stations"),
                    line_type=data.get("line_type", ""),
                    length_km=data.get("length_km"),
                    goa_level=data.get("goa_level", ""),
                    system_mode=data.get("system_mode", ""),
                    is_opened=data.get("is_opened", ""),
                    opening_date=data.get("opening_date"),
                    bid_date=data.get("bid_date"),
                    bid_amount=data.get("bid_amount"),
                    control_price=data.get("control_price"),
                    bid_link=data.get("bid_link", ""),
                    tender_link=data.get("tender_link", ""),
                    design_unit=data.get("design_unit", ""),
                    platform_software_pis=data.get("platform_software_pis", ""),
                    notes=data.get("notes", ""),
                    data_source="manual",
                )

                config = self._load_config()
                db = self._get_db(config)
                db.save_bid_raw(record)

                return jsonify({
                    "success": True,
                    "message": "中标动态记录已创建",
                    "data": {"record_id": record.record_id}
                })
            except Exception as e:
                logger.error(f"Bid dynamic create error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-dynamic/<record_id>", methods=["PUT"])
        def api_bid_dynamic_update(record_id):
            """更新中标动态记录（全字段更新）"""
            try:
                data = request.json
                if not data:
                    return jsonify({"success": False, "error": "无更新数据"}), 400

                config = self._load_config()
                db = self._get_db(config)
                data.pop("record_id", None)
                success = db.update_bid_raw(record_id, data)
                if not success:
                    return jsonify({"success": False, "error": "记录不存在"}), 404

                return jsonify({"success": True, "message": "中标动态记录已更新"})
            except Exception as e:
                logger.error(f"Bid dynamic update error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-dynamic/<record_id>", methods=["DELETE"])
        def api_bid_dynamic_delete(record_id):
            """删除中标动态记录"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                success = db.delete_bid_raw(record_id)
                if not success:
                    return jsonify({"success": False, "error": "记录不存在"}), 404

                return jsonify({"success": True, "message": "中标动态记录已删除"})
            except Exception as e:
                logger.error(f"Bid dynamic delete error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-dynamic/stats")
        def api_bid_dynamic_stats():
            """中标动态统计概要"""
            try:
                config = self._load_config()
                db = self._get_db(config)
                stats = db.get_bid_raw_stats()
                return jsonify({"success": True, "data": stats})
            except Exception as e:
                logger.error(f"Bid dynamic stats error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-dynamic/transfer", methods=["POST"])
        def api_bid_dynamic_transfer():
            """将选中的中标动态记录提取到中标数据表（bid_records）

            支持两步流程：
            1. 先检查重复（force=false），返回重复记录详情
            2. 用户确认后，再次请求携带 force=true 执行写入
            """
            try:
                data = request.json
                if not data or not data.get("record_ids"):
                    return jsonify({"success": False, "error": "请选择要提取的记录"}), 400

                record_ids = data["record_ids"]
                force = data.get("force", False)

                config = self._load_config()
                db = self._get_db(config)
                result = db.transfer_bid_raw_to_records(record_ids, force=force)

                duplicates = result.get("duplicates", [])

                # 如果检测到重复且用户未确认
                if duplicates and not force:
                    return jsonify({
                        "success": False,
                        "has_duplicates": True,
                        "message": f"检测到 {len(duplicates)} 条重复记录，请确认是否覆盖",
                        "data": {
                            "duplicate_count": len(duplicates),
                            "duplicates": [
                                {
                                    "record_id": d.get("record_id", ""),
                                    "project_name": d.get("project_name", ""),
                                    "winner": d.get("winner", ""),
                                    "bid_date": d.get("bid_date", ""),
                                    "bid_amount": d.get("bid_amount"),
                                    "category": d.get("category", ""),
                                    "_existing_record": {
                                        "project_name": d.get("_existing_record", {}).get("project_name", ""),
                                        "winner": d.get("_existing_record", {}).get("winner", ""),
                                        "bid_date": d.get("_existing_record", {}).get("bid_date", ""),
                                        "bid_amount": d.get("_existing_record", {}).get("bid_amount"),
                                    }
                                }
                                for d in duplicates
                            ]
                        }
                    })

                # 无重复，或用户已确认强制覆盖
                return jsonify({
                    "success": True,
                    "message": (
                        f"提取完成: {result['transferred']} 条已提取, "
                        f"{result['skipped']} 条跳过"
                    ),
                    "data": result
                })
            except Exception as e:
                logger.error(f"Bid dynamic transfer error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

        @app.route("/api/bid-dynamic/check-duplicate", methods=["POST"])
        def api_bid_dynamic_check_duplicate():
            """检查 bid_raw 中的记录是否在 bid_records 中已存在"""
            try:
                data = request.json
                record_ids = data.get("record_ids", [])
                if not record_ids:
                    return jsonify({"success": True, "data": {"existing": [], "new": []}})

                config = self._load_config()
                db = self._get_db(config)
                existing_in_records = set()
                with db._get_conn() as conn:
                    for rid in record_ids:
                        row = conn.execute(
                            "SELECT record_id FROM bid_records WHERE record_id = ?", (rid,)
                        ).fetchone()
                        if row:
                            existing_in_records.add(rid)

                return jsonify({
                    "success": True,
                    "data": {
                        "existing": list(existing_in_records),
                        "new": [rid for rid in record_ids if rid not in existing_in_records]
                    }
                })
            except Exception as e:
                logger.error(f"Bid dynamic check duplicate error: {e}", exc_info=True)
                return jsonify({"success": False, "error": str(e)}), 500

    def run(self, debug: bool = False):
        """启动 Web 服务器"""
        logger.info(f"RailBuddy Web 管理界面启动: http://localhost:{self.port}")
        self.app.run(
            host=self.host,
            port=self.port,
            debug=debug,
            use_reloader=False
        )
